"""
ìœ í‹¸ë¦¬í‹° í•¨ìˆ˜ (ì—‘ì…€, PDF, ë²ˆì—­, ì´ë©”ì¼ ë“±)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from googletrans import Translator
from datetime import datetime
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

def export_to_excel(data, headers, filename):
    """ë°ì´í„°ë¥¼ ì—‘ì…€ íŒŒì¼ë¡œ ë‚´ë³´ë‚´ê¸°"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # í—¤ë” ìŠ¤íƒ€ì¼
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # í—¤ë” ì‘ì„±
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # ë°ì´í„° ì‘ì„±
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # ì—´ ë„ˆë¹„ ìë™ ì¡°ì •
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    wb.save(filename)
    return filename

def export_to_pdf(data, headers, filename, title="ìœ„ì¹´ì•„ë¼ì´ ì§„ë‹¨ì‹œìŠ¤í…œ"):
    """ë°ì´í„°ë¥¼ PDF íŒŒì¼ë¡œ ë‚´ë³´ë‚´ê¸°"""
    doc = SimpleDocTemplate(filename, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#366092'),
        alignment=TA_CENTER,
        spaceAfter=30
    )
    
    # ì œëª©
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # í…Œì´ë¸” ë°ì´í„° ì¤€ë¹„
    table_data = [headers] + data
    
    # í…Œì´ë¸” ìƒì„±
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    return filename

def translate_to_japanese(text):
    """í…ìŠ¤íŠ¸ë¥¼ ì¼ë³¸ì–´ë¡œ ë²ˆì—­"""
    try:
        translator = Translator()
        result = translator.translate(text, src='ko', dest='ja')
        return result.text
    except Exception as e:
        print(f"ë²ˆì—­ ì˜¤ë¥˜: {e}")
        return text

def format_datetime(dt):
    """ë‚ ì§œì‹œê°„ í¬ë§·íŒ…"""
    if dt is None:
        return ""
    if isinstance(dt, str):
        return dt
    return dt.strftime('%Y-%m-%d %H:%M:%S')

def format_date(d):
    """ë‚ ì§œ í¬ë§·íŒ…"""
    if d is None:
        return ""
    if isinstance(d, str):
        return d
    return d.strftime('%Y-%m-%d')

def send_email(to_email, subject, body_html, body_text=None):
    """ì´ë©”ì¼ ì „ì†¡"""
    try:
        smtp_server = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.environ.get('SMTP_PORT', '587'))
        smtp_user = os.environ.get('SMTP_USER', '')
        smtp_password = os.environ.get('SMTP_PASSWORD', '')
        
        # í…ŒìŠ¤íŠ¸ ëª¨ë“œ: ì‹¤ì œ ì´ë©”ì¼ ì „ì†¡ ëŒ€ì‹  ë¡œê·¸ë§Œ ì¶œë ¥
        test_mode = os.environ.get('EMAIL_TEST_MODE', 'true').lower() == 'true'
        
        if test_mode:
            print("=" * 50)
            print("ğŸ“§ ì´ë©”ì¼ ì „ì†¡ (í…ŒìŠ¤íŠ¸ ëª¨ë“œ)")
            print(f"ìˆ˜ì‹ ì: {to_email}")
            print(f"ì œëª©: {subject}")
            print(f"ë°œì‹ ì: {smtp_user or 'wecarmobility@example.com'}")
            print("-" * 30)
            print("ë‚´ìš©:")
            # HTML íƒœê·¸ ì œê±°í•˜ì—¬ ê°„ë‹¨íˆ í‘œì‹œ
            import re
            clean_text = re.sub('<[^<]+?>', '', body_html)
            print(clean_text[:500] + "..." if len(clean_text) > 500 else clean_text)
            print("=" * 50)
            return True
        
        if not smtp_user or not smtp_password:
            error_msg = "ì´ë©”ì¼ ì„¤ì •ì´ ì—†ìŠµë‹ˆë‹¤. í™˜ê²½ë³€ìˆ˜ SMTP_USER, SMTP_PASSWORDë¥¼ ì„¤ì •í•´ì£¼ì„¸ìš”."
            print(error_msg)
            raise ValueError(error_msg)
        
        if not to_email:
            error_msg = "ìˆ˜ì‹ ì ì´ë©”ì¼ ì£¼ì†Œê°€ ì—†ìŠµë‹ˆë‹¤."
            print(error_msg)
            raise ValueError(error_msg)
        
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = smtp_user
        msg['To'] = to_email
        
        if body_text:
            part1 = MIMEText(body_text, 'plain', 'utf-8')
            msg.attach(part1)
        
        part2 = MIMEText(body_html, 'html', 'utf-8')
        msg.attach(part2)
        
        with smtplib.SMTP(smtp_server, smtp_port, timeout=30) as server:
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(msg)
        
        print(f"ì´ë©”ì¼ ì „ì†¡ ì„±ê³µ: {to_email}")
        return True
    except smtplib.SMTPAuthenticationError as e:
        error_msg = f"ì´ë©”ì¼ ì¸ì¦ ì‹¤íŒ¨: {str(e)}"
        print(error_msg)
        raise Exception(error_msg)
    except smtplib.SMTPException as e:
        error_msg = f"SMTP ì˜¤ë¥˜: {str(e)}"
        print(error_msg)
        raise Exception(error_msg)
    except Exception as e:
        error_msg = f"ì´ë©”ì¼ ì „ì†¡ ì˜¤ë¥˜: {str(e)}"
        print(error_msg)
        raise Exception(error_msg)


